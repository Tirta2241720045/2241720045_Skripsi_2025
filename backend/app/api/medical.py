from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException
from sqlalchemy.orm import Session
from app.models.database import get_db, MedicalRecord, Patient, ImageQualityMetric, User
from app.utils.logger import write_log
from app.api.auth import require_staff, require_doctor, require_staff_or_doctor
import os
import time
import numpy as np
from app.core.aes_handler import AESHandler
from app.core.lsb_handler import LSBHandler
from PIL import Image
import io
import unicodedata
from typing import Optional, Tuple
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/medical", tags=["Medical"])

AES_KEY = os.getenv("AES_KEY", "SECRET_KEY_STEGOSHIELD_2026")
aes_handler = AESHandler(AES_KEY)

DIR_ORIGINAL = os.path.join("files", "original")
DIR_EMBEDDING = os.path.join("files", "embedding")
DIR_EXTRACT = os.path.join("files", "extraction")
DIR_VISUAL = os.path.join("files", "visualization")
DIR_CSV = os.path.join("files", "csv")

for d in [DIR_ORIGINAL, DIR_EMBEDDING, DIR_EXTRACT, DIR_VISUAL, DIR_CSV]:
    os.makedirs(d, exist_ok=True)

DOCUMENTATION_XLSX_PATH = os.path.join(DIR_CSV, "documentation.xlsx")

MRI_BORDER_RATIO = 0.15

_HEADER_FILL = PatternFill("solid", start_color="2F5496")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
_DATA_FONT = Font(name="Arial", size=10)
_CENTER = Alignment(horizontal="center", vertical="center")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_LAYER_HEADERS = [
    "timestamp", "record_id", "patient_id",
    "embed_mse", "embed_psnr", "embed_ssim",
    "embed_brisque", "embed_niqe", "embed_piqe",
    "extract_mse", "extract_psnr", "extract_ssim",
    "extract_brisque", "extract_niqe", "extract_piqe",
]

_TIMING_HEADERS = [
    "timestamp", "record_id", "patient_id",
    "mri_resolution", "photo_resolution", "txt_size_kb",
    "embed_layer1_seconds", "embed_layer2_seconds", "embed_total_seconds",
    "extract_layer1_seconds", "extract_layer2_seconds", "extract_total_seconds",
]


def _safe_float(value) -> Optional[float]:
    if value is None:
        return None
    try:
        return round(float(value), 6)
    except (TypeError, ValueError):
        return None


def _init_documentation_xlsx():
    if os.path.exists(DOCUMENTATION_XLSX_PATH):
        return
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Layer 1"
    ws1.append(_LAYER_HEADERS)
    _style_header_row(ws1, len(_LAYER_HEADERS))
    ws2 = wb.create_sheet("Layer 2")
    ws2.append(_LAYER_HEADERS)
    _style_header_row(ws2, len(_LAYER_HEADERS))
    ws3 = wb.create_sheet("Timing")
    ws3.append(_TIMING_HEADERS)
    _style_header_row(ws3, len(_TIMING_HEADERS))
    wb.save(DOCUMENTATION_XLSX_PATH)


def _style_header_row(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"


def _style_data_row(ws, row: int, col_count: int):
    alt_fill = PatternFill("solid", start_color="EEF2F9") if row % 2 == 0 else None
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _DATA_FONT
        cell.border = _BORDER
        cell.alignment = _CENTER
        if alt_fill:
            cell.fill = alt_fill


def _append_layer_row(ws, ts: str, record_id: int, patient_id: int,
                      embed_metrics: dict, extract_metrics: Optional[dict]):
    row_data = [
        ts, record_id, patient_id,
        _safe_float(embed_metrics.get("mse")),
        _safe_float(embed_metrics.get("psnr")),
        _safe_float(embed_metrics.get("ssim")),
        _safe_float(embed_metrics.get("brisque")),
        _safe_float(embed_metrics.get("niqe")),
        _safe_float(embed_metrics.get("piqe")),
        _safe_float(extract_metrics.get("mse"))     if extract_metrics else None,
        _safe_float(extract_metrics.get("psnr"))    if extract_metrics else None,
        _safe_float(extract_metrics.get("ssim"))    if extract_metrics else None,
        _safe_float(extract_metrics.get("brisque")) if extract_metrics else None,
        _safe_float(extract_metrics.get("niqe"))    if extract_metrics else None,
        _safe_float(extract_metrics.get("piqe"))    if extract_metrics else None,
    ]
    ws.append(row_data)
    _style_data_row(ws, ws.max_row, len(_LAYER_HEADERS))


def _write_embed_to_xlsx(
    record_id: int,
    patient_id: int,
    mri_resolution: str,
    photo_resolution: str,
    txt_size_kb: float,
    embed_layer1_seconds: float,
    embed_layer2_seconds: float,
    embed_total_seconds: float,
    metrics_l1_embed: dict,
    metrics_l2_embed: dict,
):
    _init_documentation_xlsx()
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb = load_workbook(DOCUMENTATION_XLSX_PATH)
    ws1 = wb["Layer 1"]
    ws2 = wb["Layer 2"]
    ws3 = wb["Timing"]
    _append_layer_row(ws1, ts, record_id, patient_id, metrics_l1_embed, None)
    _append_layer_row(ws2, ts, record_id, patient_id, metrics_l2_embed, None)
    timing_row = [
        ts, record_id, patient_id,
        mri_resolution, photo_resolution, _safe_float(txt_size_kb),
        _safe_float(embed_layer1_seconds),
        _safe_float(embed_layer2_seconds),
        _safe_float(embed_total_seconds),
        None, None, None,
    ]
    ws3.append(timing_row)
    _style_data_row(ws3, ws3.max_row, len(_TIMING_HEADERS))
    wb.save(DOCUMENTATION_XLSX_PATH)


def _update_extract_to_xlsx(
    record_id: int,
    extract_layer1_seconds: float,
    extract_layer2_seconds: float,
    extract_total_seconds: float,
    metrics_l1_extract: dict,
    metrics_l2_extract: dict,
):
    _init_documentation_xlsx()
    wb = load_workbook(DOCUMENTATION_XLSX_PATH)
    ws1 = wb["Layer 1"]
    ws2 = wb["Layer 2"]
    ws3 = wb["Timing"]

    for row in ws3.iter_rows(min_row=2):
        if str(row[1].value) == str(record_id) and row[9].value is None:
            row[9].value  = _safe_float(extract_layer1_seconds)
            row[10].value = _safe_float(extract_layer2_seconds)
            row[11].value = _safe_float(extract_total_seconds)
            break

    for row in ws1.iter_rows(min_row=2):
        if str(row[1].value) == str(record_id) and row[9].value is None:
            row[9].value  = _safe_float(metrics_l1_extract.get("mse"))
            row[10].value = _safe_float(metrics_l1_extract.get("psnr"))
            row[11].value = _safe_float(metrics_l1_extract.get("ssim"))
            row[12].value = _safe_float(metrics_l1_extract.get("brisque"))
            row[13].value = _safe_float(metrics_l1_extract.get("niqe"))
            row[14].value = _safe_float(metrics_l1_extract.get("piqe"))
            break

    for row in ws2.iter_rows(min_row=2):
        if str(row[1].value) == str(record_id) and row[9].value is None:
            row[9].value  = _safe_float(metrics_l2_extract.get("mse"))
            row[10].value = _safe_float(metrics_l2_extract.get("psnr"))
            row[11].value = _safe_float(metrics_l2_extract.get("ssim"))
            row[12].value = _safe_float(metrics_l2_extract.get("brisque"))
            row[13].value = _safe_float(metrics_l2_extract.get("niqe"))
            row[14].value = _safe_float(metrics_l2_extract.get("piqe"))
            break

    wb.save(DOCUMENTATION_XLSX_PATH)


def normalize_text(text: str) -> str:
    text = text.lstrip('\ufeff')
    text = unicodedata.normalize('NFC', text)
    REPLACEMENTS = {
        '\u201c': '"', '\u201d': '"',
        '\u2018': "'", '\u2019': "'",
        '\u2014': '-', '\u2013': '-',
        '\u2012': '-', '\u2011': '-',
        '\u2010': '-', '\u2026': '...',
        '\u00a0': ' ', '\u200b': '',
        '\u200c': '', '\u200d': '',
        '\u00ad': '', '\u2002': ' ',
        '\u2003': ' ', '\u2009': ' ',
        '\u202f': ' ', '\u0000': '',
    }
    for src, dst in REPLACEMENTS.items():
        text = text.replace(src, dst)
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    return text


def normalize_text_bytes(raw: bytes) -> str:
    try:
        text = raw.decode('utf-8')
    except UnicodeDecodeError:
        text = raw.decode('latin-1')
    return normalize_text(text)


def _normalize_path(path: str) -> str:
    return path.replace('\\', '/') if path else path


def _denormalize_path(path: str) -> str:
    return os.path.normpath(path) if path else path


def _safe_delete(*paths: str) -> None:
    for p in paths:
        try:
            if p and os.path.exists(p):
                os.unlink(p)
        except OSError:
            pass


def _file_size_kb(path: str) -> float:
    try:
        return round(os.path.getsize(path) / 1024, 2) if path and os.path.exists(path) else 0.0
    except OSError:
        return 0.0


def _parse_timestamp_from_stego(stego_photo_path: str) -> Optional[str]:
    try:
        basename = os.path.splitext(os.path.basename(stego_photo_path))[0]
        parts = basename.split('_', 2)
        return parts[2] if len(parts) == 3 else None
    except Exception:
        return None


def _get_original_paths(patient_id: int, stego_photo_path: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    timestamp = _parse_timestamp_from_stego(stego_photo_path)
    if not timestamp:
        return None, None, None
    prefix = f"{patient_id}_{timestamp}"
    photo_path = os.path.join(DIR_ORIGINAL, f"photo_{prefix}.png")
    mri_path = os.path.join(DIR_ORIGINAL, f"mri_{prefix}.png")
    txt_path = os.path.join(DIR_ORIGINAL, f"medical_{prefix}.txt")
    return photo_path, mri_path, txt_path


def _save_image(img: Image.Image, path: str, compress_level: int = 0) -> None:
    img.save(path, format='PNG', compress_level=compress_level)


def _pil_to_bytes(img: Image.Image) -> bytes:
    buf = io.BytesIO()
    img.save(buf, format='PNG', compress_level=1)
    return buf.getvalue()


def _estimate_png_size(img: Image.Image) -> int:
    arr = np.array(img)
    return int(arr.size * 1.05) + 1024


def _pack_encrypted(encrypted: dict) -> bytes:
    return f"{encrypted['ciphertext']}::{encrypted['iv']}::{encrypted['mac']}".encode('utf-8')


def _unpack_encrypted(raw: str) -> Tuple[str, str, str]:
    parts = raw.split("::", 2)
    if len(parts) != 3:
        raise ValueError("Format data tidak valid: diharapkan ciphertext::iv::mac")
    return parts[0], parts[1], parts[2]


def _fmt_metrics(m) -> Optional[dict]:
    if not m:
        return None
    return {
        "layer1_mri_stego": {
            "mse": m.layer1_mse,
            "psnr": m.layer1_psnr,
            "ssim": m.layer1_ssim,
            "brisque": m.layer1_brisque,
            "niqe": m.layer1_niqe,
            "piqe": m.layer1_piqe,
        },
        "layer2_photo_stego": {
            "mse": m.layer2_mse,
            "psnr": m.layer2_psnr,
            "ssim": m.layer2_ssim,
            "brisque": m.layer2_brisque,
            "niqe": m.layer2_niqe,
            "piqe": m.layer2_piqe,
        },
        "acc_txt": {
            "acc_txt": m.acc_txt,
            "D": m.acc_txt_D,
            "T": m.acc_txt_T,
            "bit_errors": m.acc_txt_errors,
        } if m.acc_txt is not None else None,
    }


def _calculate_acctxt(original_text: str, recovered_text: str) -> dict:
    original_text = normalize_text(original_text)
    recovered_text = normalize_text(recovered_text)
    orig_bytes = original_text.encode('utf-8')
    recv_bytes = recovered_text.encode('utf-8')
    T = len(orig_bytes) * 8
    if T == 0:
        return {"acc_txt": 100.0, "D": 0, "T": 0, "bit_errors": 0}
    min_len = min(len(orig_bytes), len(recv_bytes))
    orig_bits = np.unpackbits(np.frombuffer(orig_bytes[:min_len], dtype=np.uint8))
    recv_bits = np.unpackbits(np.frombuffer(recv_bytes[:min_len], dtype=np.uint8))
    matched = int(np.sum(orig_bits == recv_bits))
    D = matched
    bit_errors = T - D
    acc_txt = round((D / T) * 100, 4)
    return {"acc_txt": acc_txt, "D": D, "T": T, "bit_errors": bit_errors}


@router.post("/upload")
async def upload_medical_data(
    patient_id: int = Form(...),
    medical_data: UploadFile = File(...),
    mri_image: UploadFile = File(...),
    patient_photo: UploadFile = File(...),
    current_user: User = Depends(require_staff),
    db: Session = Depends(get_db)
):
    patient = db.query(Patient).filter(Patient.patient_id == patient_id).first()
    if not patient:
        write_log(db, current_user.user_id, f"ERROR|UPLOAD_MEDICAL_FAILED: patient_id={patient_id}, reason=patient_not_found")
        raise HTTPException(status_code=404, detail="Pasien tidak ditemukan")

    existing_count = db.query(MedicalRecord).filter(MedicalRecord.patient_id == patient_id).count()
    if existing_count >= 10:
        write_log(db, current_user.user_id, f"ERROR|UPLOAD_MEDICAL_FAILED: patient_id={patient_id}, reason=max_records_reached")
        raise HTTPException(status_code=400, detail="Maksimal 10 rekam medis per pasien telah tercapai")

    mri_bytes = await mri_image.read()
    photo_bytes = await patient_photo.read()
    txt_bytes = await medical_data.read()

    if len(txt_bytes) > 500 * 1024:
        raise HTTPException(status_code=400, detail="Ukuran data medis maksimal 500 KB")

    txt_content = normalize_text_bytes(txt_bytes)

    try:
        img_mri = Image.open(io.BytesIO(mri_bytes))
        img_photo = Image.open(io.BytesIO(photo_bytes))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File gambar tidak valid: {str(e)}")

    mri_w, mri_h = img_mri.size
    photo_w, photo_h = img_photo.size

    if mri_w > photo_w or mri_h > photo_h:
        raise HTTPException(status_code=400, detail=f"Ukuran MRI ({mri_w}x{mri_h}) tidak boleh lebih besar dari foto pasien ({photo_w}x{photo_h}).")

    timestamp = int(time.time() * 1000)
    prefix = f"{patient_id}_{timestamp}"

    orig_photo_path = os.path.join(DIR_ORIGINAL, f"photo_{prefix}.png")
    orig_mri_path = os.path.join(DIR_ORIGINAL, f"mri_{prefix}.png")
    orig_txt_path = os.path.join(DIR_ORIGINAL, f"medical_{prefix}.txt")
    stego_out_path = os.path.join(DIR_EMBEDDING, f"stego_{prefix}.png")

    img_mri_gray = img_mri.convert('L')
    img_photo_rgb = img_photo.convert('RGB')

    try:
        img_photo_rgb.save(orig_photo_path, format='PNG', compress_level=9)
        img_mri_gray.save(orig_mri_path, format='PNG', compress_level=0)

        with open(orig_txt_path, "w", encoding="utf-8", newline='\n') as f:
            f.write(txt_content)

        encrypted = aes_handler.encrypt(txt_content)
        data_to_embed = _pack_encrypted(encrypted)

        roni_mri_capacity = LSBHandler.get_roni_capacity_border(mri_h, mri_w, MRI_BORDER_RATIO)
        roni_mri_bytes = (roni_mri_capacity // 8) - 4

        if len(data_to_embed) > roni_mri_bytes:
            raise HTTPException(status_code=400, detail=f"Data terlalu besar. Kapasitas RONI MRI: {roni_mri_bytes} bytes.")

        t1_start = time.perf_counter()
        mri_stego_img = LSBHandler.embed_to_grayscale_geometric(img_mri_gray, data_to_embed, border_ratio=MRI_BORDER_RATIO)
        time_layer1 = round(time.perf_counter() - t1_start, 6)

        mri_stego_bytes = _pil_to_bytes(mri_stego_img)
        photo_full_capacity = (photo_h * photo_w * 3 // 8) - 4

        if len(mri_stego_bytes) > photo_full_capacity:
            raise HTTPException(status_code=400, detail=f"MRI stego terlalu besar. Kapasitas foto: {photo_full_capacity} bytes.")

        t2_start = time.perf_counter()
        stego_img = LSBHandler.embed_to_rgb_full(img_photo_rgb, mri_stego_img)
        time_layer2 = round(time.perf_counter() - t2_start, 6)

        time_embed_total = round(time_layer1 + time_layer2, 6)

        stego_img.save(stego_out_path, format='PNG', compress_level=9)

        metrics_l1 = LSBHandler.calculate_metrics(img_mri_gray, mri_stego_img, mode='L')
        metrics_l2 = LSBHandler.calculate_metrics(img_photo_rgb, stego_img, mode='RGB')
        nriqa_l1 = LSBHandler.calculate_nriqa_metrics(mri_stego_img, mode='L')
        nriqa_l2 = LSBHandler.calculate_nriqa_metrics(stego_img, mode='RGB')

        file_sizes = {
            "original_txt_kb": _file_size_kb(orig_txt_path),
            "original_mri_kb": _file_size_kb(orig_mri_path),
            "original_photo_kb": _file_size_kb(orig_photo_path),
            "stego_kb": _file_size_kb(stego_out_path),
        }

        db_record = MedicalRecord(
            patient_id=patient_id,
            medical_data_path=_normalize_path(orig_txt_path),
            photo_path=_normalize_path(orig_photo_path),
            mri_path=_normalize_path(orig_mri_path),
            stego_photo_path=_normalize_path(stego_out_path),
            embed_time_seconds=time_embed_total,
        )
        db.add(db_record)
        db.commit()
        db.refresh(db_record)

        db.add(ImageQualityMetric(
            record_id=db_record.record_id,
            layer1_mse=metrics_l1['mse'],
            layer1_psnr=metrics_l1['psnr'],
            layer1_ssim=metrics_l1['ssim'],
            layer1_brisque=nriqa_l1['brisque'],
            layer1_niqe=nriqa_l1['niqe'],
            layer1_piqe=nriqa_l1['piqe'],
            layer2_mse=metrics_l2['mse'],
            layer2_psnr=metrics_l2['psnr'],
            layer2_ssim=metrics_l2['ssim'],
            layer2_brisque=nriqa_l2['brisque'],
            layer2_niqe=nriqa_l2['niqe'],
            layer2_piqe=nriqa_l2['piqe'],
        ))
        db.commit()

        _write_embed_to_xlsx(
            record_id=db_record.record_id,
            patient_id=patient_id,
            mri_resolution=f"{mri_w}x{mri_h}",
            photo_resolution=f"{photo_w}x{photo_h}",
            txt_size_kb=file_sizes["original_txt_kb"],
            embed_layer1_seconds=time_layer1,
            embed_layer2_seconds=time_layer2,
            embed_total_seconds=time_embed_total,
            metrics_l1_embed={**metrics_l1, **nriqa_l1},
            metrics_l2_embed={**metrics_l2, **nriqa_l2},
        )

        write_log(db, current_user.user_id, f"UPLOAD_MEDICAL: patient_id={patient_id}, record_id={db_record.record_id}")

        return {
            "message": "Data berhasil diproses",
            "record_id": db_record.record_id,
            "stego_image": _normalize_path(stego_out_path),
            "roni_type": "geometric_border_layer1_only",
            "embed_time": {
                "layer1_seconds": time_layer1,
                "layer2_seconds": time_layer2,
                "total_seconds": time_embed_total,
            },
            "quality_metrics": {
                "layer1_mri_stego": {**metrics_l1, **nriqa_l1},
                "layer2_photo_stego": {**metrics_l2, **nriqa_l2},
            },
            "file_sizes": file_sizes,
        }

    except HTTPException:
        raise
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        write_log(db, current_user.user_id, f"ERROR|UPLOAD_MEDICAL_UNEXPECTED: {e}")
        raise HTTPException(status_code=500, detail=f"Terjadi kesalahan: {str(e)}")


@router.get("/patient/{patient_id}")
async def get_medical_records_by_patient(
    patient_id: int,
    current_user: User = Depends(require_staff_or_doctor),
    db: Session = Depends(get_db)
):
    patient = db.query(Patient).filter(Patient.patient_id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Pasien tidak ditemukan")

    records = db.query(MedicalRecord).filter(MedicalRecord.patient_id == patient_id).order_by(MedicalRecord.record_id.desc()).all()

    result = []
    for record in records:
        all_metrics = db.query(ImageQualityMetric).filter(ImageQualityMetric.record_id == record.record_id).order_by(ImageQualityMetric.metric_id.asc()).all()
        embed_m = all_metrics[0] if len(all_metrics) > 0 else None
        extract_m = all_metrics[1] if len(all_metrics) > 1 else None

        orig_photo_path, orig_mri_path, orig_txt_path = _get_original_paths(record.patient_id, record.stego_photo_path)
        timestamp = _parse_timestamp_from_stego(record.stego_photo_path)
        prefix = f"{record.patient_id}_{timestamp}" if timestamp else None
        vis_mri_path = os.path.join(DIR_VISUAL, f"vis_mri_{prefix}.png") if prefix else None
        vis_photo_path = os.path.join(DIR_VISUAL, f"vis_photo_{prefix}.png") if prefix else None

        file_sizes = {
            "original_txt_kb": _file_size_kb(orig_txt_path) if orig_txt_path else 0.0,
            "original_mri_kb": _file_size_kb(orig_mri_path) if orig_mri_path else 0.0,
            "original_photo_kb": _file_size_kb(orig_photo_path) if orig_photo_path else 0.0,
            "stego_kb": _file_size_kb(_denormalize_path(record.stego_photo_path)),
            "vis_mri_kb": _file_size_kb(vis_mri_path) if vis_mri_path else 0.0,
            "vis_photo_kb": _file_size_kb(vis_photo_path) if vis_photo_path else 0.0,
        }

        result.append({
            "record_id": record.record_id,
            "medical_data_path": record.medical_data_path,
            "photo_path": record.photo_path,
            "mri_path": record.mri_path,
            "stego_photo_path": record.stego_photo_path,
            "roni_type": "geometric_border_layer1_only",
            "visualization": {
                "mri_lsb_map": _normalize_path(vis_mri_path) if vis_mri_path and os.path.exists(vis_mri_path) else None,
                "photo_lsb_map": _normalize_path(vis_photo_path) if vis_photo_path and os.path.exists(vis_photo_path) else None,
            },
            "upload_date": record.created_at.isoformat() if record.created_at else None,
            "quality_metrics": {
                "embedding": _fmt_metrics(embed_m),
                "extraction": _fmt_metrics(extract_m),
            },
            "file_sizes": file_sizes,
        })

    return {
        "patient_id": patient_id,
        "patient_name": patient.full_name,
        "total_records": len(records),
        "records": result,
    }


@router.get("/extract/{record_id}")
async def extract_medical_data(
    record_id: int,
    current_user: User = Depends(require_doctor),
    db: Session = Depends(get_db)
):
    record = db.query(MedicalRecord).filter(MedicalRecord.record_id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Rekam medis tidak ditemukan")

    stego_path = _denormalize_path(record.stego_photo_path)
    if not os.path.exists(stego_path):
        raise HTTPException(status_code=404, detail="File stego tidak ditemukan")

    patient = db.query(Patient).filter(Patient.patient_id == record.patient_id).first()
    orig_photo_path, orig_mri_path, orig_txt_path = _get_original_paths(record.patient_id, record.stego_photo_path)

    timestamp = _parse_timestamp_from_stego(record.stego_photo_path)
    ext_prefix = f"{record.patient_id}_{timestamp}_{record_id}" if timestamp else f"{record.patient_id}_{record_id}"

    ext_mri_path = os.path.join(DIR_EXTRACT, f"mri_{ext_prefix}.png")
    ext_photo_path = os.path.join(DIR_EXTRACT, f"photo_{ext_prefix}.png")
    ext_txt_path = os.path.join(DIR_EXTRACT, f"medical_{ext_prefix}.txt")

    _safe_delete(ext_mri_path, ext_photo_path, ext_txt_path)

    try:
        stego_img = Image.open(stego_path).convert('RGB')

        t2_start = time.perf_counter()
        extracted_mri_img = LSBHandler.extract_from_rgb_full(stego_img)
        time_extract_layer2 = round(time.perf_counter() - t2_start, 6)

        if extracted_mri_img is None:
            raise HTTPException(status_code=500, detail="Gagal mengekstrak MRI dari stego")

        t1_start = time.perf_counter()
        extracted_bytes = LSBHandler.extract_from_grayscale_geometric(extracted_mri_img, border_ratio=MRI_BORDER_RATIO)
        time_extract_layer1 = round(time.perf_counter() - t1_start, 6)

        if not extracted_bytes:
            raise HTTPException(status_code=500, detail="Gagal menemukan data tersembunyi")

        time_extract_total = round(time_extract_layer1 + time_extract_layer2, 6)

        raw = extracted_bytes.decode("utf-8")
        ciphertext, iv, mac = _unpack_encrypted(raw)
        decrypted = aes_handler.decrypt(ciphertext, iv, mac)
        decrypted = normalize_text(decrypted)

        with open(ext_txt_path, "w", encoding="utf-8", newline='\n') as f:
            f.write(decrypted)

        _save_image(extracted_mri_img, ext_mri_path, compress_level=0)

        stego_array = np.array(stego_img, dtype=np.uint8)
        cleaned_photo_array = stego_array & np.uint8(0xFE)
        cleaned_photo_img = Image.fromarray(cleaned_photo_array, mode='RGB')
        _save_image(cleaned_photo_img, ext_photo_path, compress_level=9)

        acctxt_result = {"acc_txt": None, "D": None, "T": None, "bit_errors": None}
        if orig_txt_path and os.path.exists(orig_txt_path):
            with open(orig_txt_path, "r", encoding="utf-8", newline='') as f:
                original_text = f.read()
            acctxt_result = _calculate_acctxt(original_text, decrypted)

        metrics_l1 = {"mse": 0.0, "psnr": 100.0, "ssim": 1.0}
        if orig_mri_path and os.path.exists(orig_mri_path):
            orig_mri_img = Image.open(orig_mri_path)
            if orig_mri_img.mode != 'L':
                orig_mri_img = orig_mri_img.convert('L')
            metrics_l1 = LSBHandler.calculate_metrics(orig_mri_img, extracted_mri_img, mode='L')

        metrics_l2 = {"mse": 0.0, "psnr": 100.0, "ssim": 1.0}
        if orig_photo_path and os.path.exists(orig_photo_path):
            orig_photo_img = Image.open(orig_photo_path)
            if orig_photo_img.mode != 'RGB':
                orig_photo_img = orig_photo_img.convert('RGB')
            metrics_l2 = LSBHandler.calculate_metrics(orig_photo_img, cleaned_photo_img, mode='RGB')

        nriqa_l1 = LSBHandler.calculate_nriqa_metrics(extracted_mri_img, mode='L')
        nriqa_l2 = LSBHandler.calculate_nriqa_metrics(cleaned_photo_img, mode='RGB')

        record.extract_time_seconds = time_extract_total
        db.commit()

        _update_extract_to_xlsx(
            record_id=record_id,
            extract_layer1_seconds=time_extract_layer1,
            extract_layer2_seconds=time_extract_layer2,
            extract_total_seconds=time_extract_total,
            metrics_l1_extract={**metrics_l1, **nriqa_l1},
            metrics_l2_extract={**metrics_l2, **nriqa_l2},
        )

        all_metrics = db.query(ImageQualityMetric).filter(ImageQualityMetric.record_id == record_id).order_by(ImageQualityMetric.metric_id.asc()).all()
        if len(all_metrics) >= 2:
            m = all_metrics[1]
            m.layer1_mse = metrics_l1['mse']
            m.layer1_psnr = metrics_l1['psnr']
            m.layer1_ssim = metrics_l1['ssim']
            m.layer1_brisque = nriqa_l1['brisque']
            m.layer1_niqe = nriqa_l1['niqe']
            m.layer1_piqe = nriqa_l1['piqe']
            m.layer2_mse = metrics_l2['mse']
            m.layer2_psnr = metrics_l2['psnr']
            m.layer2_ssim = metrics_l2['ssim']
            m.layer2_brisque = nriqa_l2['brisque']
            m.layer2_niqe = nriqa_l2['niqe']
            m.layer2_piqe = nriqa_l2['piqe']
            m.acc_txt = acctxt_result['acc_txt']
            m.acc_txt_D = acctxt_result['D']
            m.acc_txt_T = acctxt_result['T']
            m.acc_txt_errors = acctxt_result['bit_errors']
        else:
            db.add(ImageQualityMetric(
                record_id=record_id,
                layer1_mse=metrics_l1['mse'],
                layer1_psnr=metrics_l1['psnr'],
                layer1_ssim=metrics_l1['ssim'],
                layer1_brisque=nriqa_l1['brisque'],
                layer1_niqe=nriqa_l1['niqe'],
                layer1_piqe=nriqa_l1['piqe'],
                layer2_mse=metrics_l2['mse'],
                layer2_psnr=metrics_l2['psnr'],
                layer2_ssim=metrics_l2['ssim'],
                layer2_brisque=nriqa_l2['brisque'],
                layer2_niqe=nriqa_l2['niqe'],
                layer2_piqe=nriqa_l2['piqe'],
                acc_txt=acctxt_result['acc_txt'],
                acc_txt_D=acctxt_result['D'],
                acc_txt_T=acctxt_result['T'],
                acc_txt_errors=acctxt_result['bit_errors'],
            ))
        db.commit()

        write_log(db, current_user.user_id, f"EXTRACT_MEDICAL: record_id={record_id}, patient_id={record.patient_id}")

        return {
            "record_id": record_id,
            "patient_id": record.patient_id,
            "patient_name": patient.full_name if patient else "Unknown",
            "medical_data": decrypted,
            "extract_time_seconds": time_extract_total,
            "extract_time_per_layer": {
                "layer1_seconds": time_extract_layer1,
                "layer2_seconds": time_extract_layer2,
                "total_seconds": time_extract_total,
            },
            "stego_image": record.stego_photo_path,
            "photo_path": _normalize_path(ext_photo_path),
            "mri_path": _normalize_path(ext_mri_path),
            "txt_path": _normalize_path(ext_txt_path),
            "lsb_extraction_success": True,
            "roni_type": "geometric_border_layer1_only",
            "acc_txt": acctxt_result,
            "quality_metrics": {
                "extraction": {
                    "layer1_mri_stego": {**metrics_l1, **nriqa_l1},
                    "layer2_photo_stego": {**metrics_l2, **nriqa_l2},
                    "acc_txt": acctxt_result,
                }
            },
            "file_sizes": {
                "stego_kb": _file_size_kb(stego_path),
                "extracted_mri_kb": _file_size_kb(ext_mri_path),
                "extracted_photo_kb": _file_size_kb(ext_photo_path),
                "extracted_txt_kb": _file_size_kb(ext_txt_path),
            },
        }

    except HTTPException:
        _safe_delete(ext_mri_path, ext_photo_path, ext_txt_path)
        raise
    except ValueError as e:
        _safe_delete(ext_mri_path, ext_photo_path, ext_txt_path)
        raise HTTPException(status_code=422, detail=str(e))
    except Exception as e:
        write_log(db, current_user.user_id, f"ERROR|EXTRACT_MEDICAL_UNEXPECTED: {e}")
        _safe_delete(ext_mri_path, ext_photo_path, ext_txt_path)
        raise HTTPException(status_code=500, detail=f"Terjadi kesalahan: {str(e)}")


@router.delete("/record/{record_id}")
async def delete_medical_record(
    record_id: int,
    current_user: User = Depends(require_staff),
    db: Session = Depends(get_db)
):
    record = db.query(MedicalRecord).filter(MedicalRecord.record_id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Rekam medis tidak ditemukan")

    patient_id = record.patient_id
    db.query(ImageQualityMetric).filter(ImageQualityMetric.record_id == record_id).delete()

    timestamp = _parse_timestamp_from_stego(record.stego_photo_path)
    prefix = f"{record.patient_id}_{timestamp}" if timestamp else None
    ext_prefix = f"{record.patient_id}_{timestamp}_{record_id}" if timestamp else f"{record.patient_id}_{record_id}"

    all_paths = [
        record.medical_data_path, record.photo_path, record.mri_path, record.stego_photo_path,
    ]

    if prefix:
        all_paths += [
            os.path.join(DIR_ORIGINAL, f"photo_{prefix}.png"),
            os.path.join(DIR_ORIGINAL, f"mri_{prefix}.png"),
            os.path.join(DIR_ORIGINAL, f"medical_{prefix}.txt"),
            os.path.join(DIR_VISUAL, f"vis_mri_{prefix}.png"),
            os.path.join(DIR_VISUAL, f"vis_photo_{prefix}.png"),
        ]

    all_paths += [
        os.path.join(DIR_EXTRACT, f"mri_{ext_prefix}.png"),
        os.path.join(DIR_EXTRACT, f"photo_{ext_prefix}.png"),
        os.path.join(DIR_EXTRACT, f"medical_{ext_prefix}.txt"),
    ]

    seen = set()
    deleted = []
    for path in all_paths:
        if not path:
            continue
        real = _denormalize_path(path)
        if real in seen:
            continue
        seen.add(real)
        if os.path.exists(real):
            try:
                os.unlink(real)
                deleted.append(_normalize_path(real))
            except OSError:
                pass

    db.delete(record)
    db.commit()
    write_log(db, current_user.user_id, f"DELETE_MEDICAL_RECORD: record_id={record_id}, patient_id={patient_id}")
    return {"message": f"Rekam medis #{record_id} berhasil dihapus", "record_id": record_id, "files_deleted": {"deleted": deleted, "count": len(deleted)}}